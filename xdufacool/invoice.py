# invoice.py ---
#
# Filename: invoice.py
# Author: Fred Qi
# Created: 2022-09-01 13:12:59(+0800)
#
# Last-Updated: 2023-03-23 18:50:50(+0800) [by Fred Qi]
#     Update #: 346
# 

# Commentary:
#
#
# 

# Change Log:
#
#
# 

import re
import fitz
from os import path
from glob import glob
from openpyxl import Workbook
from argparse import ArgumentParser


INVOICE_PATH = "/home/fred/cloud/OneDrive/office/invoices/221109"


class InvoiceParser(object):

    def __init__(self, filename):
        invoice = fitz.open(filename)
        page = invoice.load_page(0)
        self.words = page.get_text('words')
        self.fields = set(["发票代码", "发票号码", "开票日期", "小写"])
        self.word_dict = self.word_backref(self.words, self.fields)

    @staticmethod
    def word_backref(words, fields):
        word_dict = {}
        for word in words:
            text = re.sub(r"[:\(\)]+", "", word[4])
            if text in fields:
                word_dict[text] = word
                fields -= set([text])
        return word_dict

    @staticmethod
    def find_right_word(words, word_ref):
        # print(word_ref)
        l_ref, t_ref, r_ref, b_ref = word_ref[:4]
        y_ref = 0.5*(t_ref + b_ref)
        right_words = []
        for word in words:            
            l, t, r, b = word[:4]
            y = 0.5*(t+b)
            if abs(y-y_ref) < 5.0 and 0 < l - r_ref < 120:
                # print(" ", word)
                right_words.append(word)
        return sorted(right_words, key=lambda rc: rc[0])
        
    def parse_code(self):
        self.form = {}
        for key, value in self.word_dict.items():
            words = self.find_right_word(self.words, value)
            self.form[key] = "".join([word[4] for word in words])


def merge_invoices(invoices, invoice_merged):
    """Merge digital invoices given by invoices in to one PDF file invoice_merged."""
    output = fitz.open()

    for pdf in invoices:
        with fitz.open(pdf) as invoice:
            output.insert_pdf(invoice)

    output.save(invoice_merged)
    output.close()
    

def parsing_jd(invoices, statfile):

    cols = ["物品", "发票代码", "发票号码", "开票日期", "小写", "报销"]

    wb = Workbook()
    ws = wb.active
    ws.title = "财务"
    ws.append(cols)
    
    # INVOICE_FILENAMES = glob(path.join(INVOICE_PATH, "京东发票-*.pdf"))
    header = False
    for filepath in invoices:
        filename = path.basename(filepath)             
        parser = InvoiceParser(filepath)
        parser.parse_code()
        values  = [filename.split("-")[1]]
        values += [parser.form[col] for col in cols[1:-2]]
        values += [float(parser.form[cols[-2]].replace("¥", ""))]
        values += [1]
        ws.append(values)

    n_lines = len(invoices) + 1

    ws[f"E{n_lines+1}"] = f"=SUMIF(F2:F{n_lines},1,E2:E{n_lines})"
    # wb.save("invoices.xlsx")
    basefile, _ = path.splitext(statfile)
    wb.save(basefile + ".xlsx")
    # print(filename)
    # print(" ", parser.form)
        
    # doc = fitz.open(filename)
    # page = doc.load_page(0)
    # # print(page.links())
    # blocks = page.get_text("blocks")
    # for blk in blocks:
    #     text = blk[4]
    #     if text.find("合") >= 0:
    #         print(blk)

    # reader = PdfReader(filename)
    # page = reader.getPage(0)
    # item = page.get_contents()[0]
    # # help(item)
    # obj = item.get_object()
    # print(obj.values())
    # print(obj.flate_encode())
    # print(type(item), item, type(obj), obj)

    # help(PyPDF2.generic._data_structures.EncodedStreamObject)
    # page = reader.pages[0]
    # text = page.extract_text()
    # print(text)

    
def collect_invoice():
    desc = "To collect digital invoices in PDF format."
    parser = ArgumentParser(description=desc)
    parser.add_argument('-o', '--output', type=str, dest='invoices_merged',
                        help="The merged PDF containing all given digital invoices.")
    parser.add_argument('-s', '--statistics', type=str, dest='invoices_stat',
                        help="Generate statistics of JD invoices.")
    parser.add_argument('invoices', type=str, nargs="+",
                        help="Invoices to be collected.")
    args = parser.parse_args()

    # print(args)
    if args.invoices_merged:
        merge_invoices(args.invoices, args.invoices_merged)
    if args.invoices_stat:
        parsing_jd(args.invoices, args.invoices_stat)
        
    

if "__main__" == __name__:
    desc = "To collect digital invoices in PDF format."
    parser = ArgumentParser(description=desc)
    parser.add_argument('-o', '--output', type=str, dest='invoices_merged',
                        help="The merged PDF containing all given digital invoices.")
    parser.add_argument('invoices', type=str, nargs="+",
                        help="Invoices to be collected.")
    args = parser.parse_args()

    # print(args)
    if args.invoices_merged:
        merge_invoices(args.invoices, args.invoices_merged)        

    # cols = ["物品", "发票代码", "发票号码", "开票日期", "小写", "报销"]
    # for filepath in args.invoices:
    #     filename = path.basename(filepath)             
    #     invoice_parser = InvoiceParser(filepath)
    #     invoice_parser.parse_code()
    #     print(invoice_parser.form)
    #     values  = [filename.split("-")[1]]
    #     values += [invoice_parser.form[col] for col in cols[1:-2]]
    #     values += [float(invoice_parser.form[cols[-2]].replace("¥", ""))]
    #     values += [1]
    #     ws.append(values)

    # n_lines = len(args.invoices) + 1

    # ws[f"E{n_lines+1}"] = f"=SUMIF(F2:F{n_lines},1,E2:E{n_lines})"
    # wb.save("invoices.xlsx")

# 
# invoice.py ends here
