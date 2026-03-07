# score_helper.py ---
#
# Filename: score_helper.py
# Author: Fred Qi
# Created: 2021-01-10 20:41:42(+0800)
#
# Last-Updated: 2023-06-25 16:36:01(+0800) [by Fred Qi]
#     Update #: 988
# 

# Commentary:
#
#
# 

# Change Log:
#
#
#
import os
import csv
import sys
import time
# import xlrd
# import xlwt
import openpyxl
import numpy as np
from itertools import groupby
from dataclasses import dataclass, field

import shutil
# import click


@dataclass
class ScoreInterval(object):

    left: float
    right: float
    desc: str
    count: int = field(init=False)
    percent: float = field(init=False)

    def __str__(self):
        # return f'{self.desc:4} ({self.left:4.1f}-{self.right:5.1f}): {self.count:d}({self.percent:5.2f}%)'
        return f'{self.desc}: {self.count:d} ({self.percent:5.2f}%)'

    def __post_init__(self):
        self.count = 0
        self.percent = 0.0

    def grading(self, scores):
        left_cond = scores >= self.left
        if self.right < 100:            
            right_cond = scores < self.right
        else:
            right_cond = scores <= self.right
        indices = np.logical_and(left_cond, right_cond)
        self.count = np.count_nonzero(indices)
        self.percent = self.count*100.0/scores.size if scores.size > 0 else 0.0


class ScoreStat(object):
    def __init__(self, scores):
        intervals = [(90, 100, '优秀'), (80, 90, '良好'),
                     (70, 80, '中等'), (60, 70, '及格'), (0, 60, '不及格')]
        self.stat = [ScoreInterval(*value) for value in intervals]
        self.average = 0.0
        self.n_students = 0
        self.__grading(scores)

    def __str__(self):
        lines = []
        for sc_interval in self.stat:
            lines.append(str(sc_interval))
        return f'{self.n_students}/{self.average:5.2f}\n' + '\n'.join(lines)

    def __grading(self, scores):
        scs = np.array([float(x) if x else -1 for x in scores])
        self.average = scs.mean() if scs.size > 0 else 0.0
        self.n_students = scs.size
        for sc_interval in self.stat:
            sc_interval.grading(scs)

    def get_table(self):
        """
        Returns the score statistics as a list of tuples.
        Each tuple: (level, range, count, percentage)
        """
        table = []
        for interval in self.stat:
            range_str = f"{interval.left:.0f}-{interval.right:.0f}"
            table.append((f"{interval.desc:^{7}}", f"{range_str:^{7}}",
                         f"{interval.count:^{5}d}", f"{interval.percent:>{5}.1f}%"))
        return table


class ScoreAnalysis(object):
    def __init__(self):
        self._scores = {}
        # a tuple of abilities contains subtotal, weight, prop which are:
        # - total scores of subitem
        # - weight of corresponding graduation indicator
        # - proportion of homework scores
        # and the key is the name of corresponding graduation indicator.
        self.abilities = {'ind14': (35.0, 0.1, 0.50),
                          'ind33': (45.0, 0.3, 0.60),
                          'ind42': (20.0, 0.1, 0.45)}

    def __str__(self):
        lines = []
        for key in self._scores.keys():
            lines.append(f'{key:8}: ' + self.get_text(key))
        return '\n'.join(lines)

    def _calc_abilities(self):
        sc_homeworks = self._scores['avg_homeworks']
        sc_final = self._scores['avg_final']

        total = sum([v[0] for v in self.abilities.values()])
        for key, (subtotal, weight, prop) in self.abilities.items():
            sc_item = (prop*sc_homeworks + (1-prop)*sc_final)/total
            self._scores[f'avg_{key}'] = sc_item*subtotal
            self._scores[f'percent_{key}'] = sc_item
            self._scores[f'grad_{key}'] = sc_item*weight

    def add_parts(self, keys, scores):
        for key, score in zip(keys, scores):
            score_np = np.array([float(x) if x else -1 for x in score])
            self._scores[f'avg_{key}'] = score_np.mean()
        self._calc_abilities()
    
    def get_text(self, key):
        value = self._scores.get(key, -1)
        if key.find('percent') >= 0 or key.find('grad') >= 0:
            text = f'{value*100:5.2f}%'
        else:
            text = f'{value:4.1f}'
        return text

    def get_analysis(self):
        return {key: self.get_text(key) for key in self._scores.keys()}

def normalize_header(header_row):
    """Normalize headers to string and strip whitespace."""
    return [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]

def load_source_data(filepath):
    print(f"Loading source data from: {filepath}")
    if not os.path.exists(filepath):
        print(f"Error: File '{filepath}' not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    # Get headers from first row
    headers = normalize_header(sheet[1])
    
    # Define required and optional columns
    required_cols = {'学号'}
    score_cols = ['客观分', '主观分', '成绩']
    
    # Check if '学号' exists
    if '学号' not in headers:
        print("Error: Column '学号' not found in source file.")
        sys.exit(1)
        
    col_indices = {}
    col_indices['学号'] = headers.index('学号')
    
    for col in score_cols:
        if col in headers:
            col_indices[col] = headers.index(col)
        else:
            print(f"Warning: Column '{col}' not found in source file. It will be skipped.")
    
    data = {}
    # Iterate rows starting from row 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # row is a tuple of values
        try:
            student_id = str(row[col_indices['学号']]).strip()
            if not student_id or student_id == 'None':
                continue
                
            entry = {}
            for col_name in score_cols:
                if col_name in col_indices:
                    entry[col_name] = row[col_indices[col_name]]
            
            data[student_id] = entry
        except IndexError:
            continue
            
    print(f"Loaded {len(data)} records from source file.")
    return data

def merge_to_target(target_path, data, output_path=None):
    print(f"Merging into target: {target_path}")
    if not os.path.exists(target_path):
        print(f"Error: File '{target_path}' not found.")
        sys.exit(1)

    wb = openpyxl.load_workbook(target_path)
    sheet = wb.active
    
    # Get headers from target
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    
    if '学号' not in headers:
        print("Error: Column '学号' not found in target file.")
        sys.exit(1)
    
    id_col_idx = headers.index('学号') + 1 # 1-based index for cell access
    
    # Score columns to update
    score_cols = ['客观分', '主观分', '成绩']
    col_map = {}
    
    # Find or create columns in target
    for col_name in score_cols:
        if col_name in headers:
            col_map[col_name] = headers.index(col_name) + 1
        else:
            new_col_idx = len(headers) + 1
            sheet.cell(row=1, column=new_col_idx, value=col_name)
            headers.append(col_name)
            col_map[col_name] = new_col_idx
            print(f"Added new column '{col_name}' to target file.")

    count = 0
    # Update rows
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        # Access student ID cell
        id_cell = sheet.cell(row=row_idx, column=id_col_idx)
        student_id = str(id_cell.value).strip()
        
        if student_id in data:
            record = data[student_id]
            for col_name, val in record.items():
                target_col_idx = col_map[col_name]
                sheet.cell(row=row_idx, column=target_col_idx, value=val)
            count += 1
            
    out_file = output_path if output_path else target_path
    wb.save(out_file)
    print(f"Successfully updated {count} records.")
    print(f"Saved result to: {out_file}")


def split_ranking(data, col_sub, col_rank):
    """Split scores into two categories according to whether it is ranked online."""
    def rank_key(row, col_sub=col_sub, col_rank=col_rank):
        return f"{row[col_rank]}-{row[col_sub]}"
    data_usub = filter(lambda x: "未评阅-已提交" != rank_key(x), data)
    data_rank = filter(lambda x: "未评阅-已提交" == rank_key(x), data)
    return list(data_rank), list(data_usub)
    

def normal_cumsum(n_students, score_max=98, score_min=80):
    dist = np.random.normal(size=n_students)
    n_blocks = score_max - score_min + 1
    hist, edges = np.histogram(dist, bins=n_blocks)
    cumsum = np.cumsum(hist)
    return cumsum


def time_based_rank(data, col_sc, col_time,
                    score_max=98, score_min=80):
    """Ranking homeworks based on submission time."""
    time_format = "%Y-%m-%d %H:%M:%S"
    def key_time(time_str):
        return time.strptime(time_str, time_format)
        
    data_rank = sorted(data, key=lambda row: key_time(row[col_time]))
    cumsum = normal_cumsum(len(data_rank), score_max, score_min)
    idx_cumsum, sc = 0, score_max
    for idx in range(len(data_rank)):
        while cumsum[idx_cumsum] <= idx:
            idx_cumsum += 1
            sc -= 1
        if idx < cumsum[idx_cumsum]:
            data_rank[idx][col_sc] = sc
    return data_rank
    

def find_column_index(header, fields):
    """Find the corresponding column indices of given fields."""
    indices = []
    for txt in fields:
        for idx, col in enumerate(header):
            if col.find(txt) >= 0:
                indices.append(idx)
                break
    return indices


def load_xls(xlsfile):
    book = xlrd.open_workbook(xlsfile)
    sh = book.sheet_by_index(0)
    data = []
    for rx in range(sh.nrows):
        row = sh.row(rx)
        data.append([c.value for c in row])
    return data


def value_dict(data, sid_col, value_col, convert=False):
    students = {}
    for row in data[1:]:
        sid, name = row[sid_col], row[value_col]
        value = float(name) if convert else name.strip()
        students[sid] = value
    return students


def merge_students(source_a, source_b):
    results = dict(source_a)
    for key, val in source_b.items():
        results[key] = val
    return results


def student_form(students, sid_label, name_label):
    """Convert student dict to a table."""
    data = [[sid_label, name_label]]
    for key, value in students.items():
        data.append([key, value])
    return data


def merge_scores(scores, weights=None, n_digits=0):
    s_keys = list(scores.keys())
    # headers = ["姓名", "学号"] + s_keys + ["加权平均分"]
    if weights is None:
        weights = {key: 1.0 for key in s_keys}
    w_sum = sum(list(weights.values()))
    
    students = []
    for key in s_keys:
        students += list(scores[key].keys())
    students = sorted(list(set(students)))
    
    data = {}
    errors = []
    for sid in students:
        avg = 0.0
        scs = {"学号": sid}
        for key in s_keys:
            try:
                sc = float(scores[key].get(sid, 0))
            except ValueError as e:
                sc = 0.0
                errors.append([sid, key, scores[key][sid]])
            avg += weights[key] * sc
            scs[key] = sc
        avg /= w_sum
        scs["加权平均分"] = round(avg, n_digits)
        if avg < 60:
            errors.append([sid, avg])
        data[sid] = scs
    return data, errors


def score_dict_to_table(scores, fields):
    """Convert scores to a table."""
    form = [fields]
    for sid, values in scores.items():
        row = [values[fields[0]]]
        #row += ["{v:g}".format(v=values[key]) for key in fields[1:]]
        row += [values[key] for key in fields[1:]]
        form.append(row)
    return form


def fill_score_form(form, scores, sid_col=1, score_col=-2):
    """Fill in the score form, whose first line is a header."""
    # Skipping the first row (header)
    for row in form[1:]:
        sid = row[sid_col]
        sc = scores[sid] if sid in scores else 0
        row[score_col] = f"{sc:g}"


def write_csv(data, csvfile):
    """Save 2D list (data) to a csv file."""
    with open(csvfile, "w") as csv_stream:
        writer = csv.writer(csv_stream)
        writer.writerows(data)

        
def write_xls(form, xlspath, sheet_name="sheet"):
    """Write a form to xls file for uploading."""

    book = xlwt.Workbook()
    sheet = book.add_sheet(sheet_name)
    for r, row in enumerate(form):
        for c, value in enumerate(row):
            sheet.write(r, c, value)
    book.save(xlspath)


if __name__ == "__main__":
    scores = np.array([95, 85, 75, 65, 55, 95, 85, 75, 65, 55])
    score_stat = ScoreStat(scores)
    # print(score_stat)
    table = score_stat.get_table()
    for row in table:
        print(row)

# @click.group()
# def xduscore():
#     pass


# @xduscore.command()
# @click.argument('students', nargs=-1, type=click.Path(exists=True))
# @click.option('-o', '--output', type=click.Path())
# @click.option('--name-xdu', default='姓名(文本)', type=click.STRING)
# @click.option('--sid-xdu', default='学号(文本)', type=click.STRING)
# @click.option('--name-baidu', default='学生姓名', type=click.STRING)
# @click.option('--sid-baidu', default='学生学号', type=click.STRING)
# @click.option('--sheet-baidu', default='学生名单', type=click.STRING)
# def merge(output, students, **kwargs):
#     """Read lists of students downloaded from Xidian, merge them to create a unified list."""    
#     students_all = {}
#     fields = [kwargs["name_xdu"], kwargs["sid_xdu"]]
#     for xlsfile in students:
#         data = load_xls(xlsfile)
#         name_col, sid_col = find_column_index(data[0], fields)
#         stus = value_dict(data, sid_col, name_col)
#         students_all = merge_students(students_all, stus)

#     l_sid, l_name = kwargs["sid_baidu"], kwargs["name_baidu"]
#     form = student_form(students_all, l_sid, l_name)
#     # write_csv(form, output.replace(".xls", ".csv"))
#     write_xls(form, output, kwargs["sheet_baidu"])

# @xduscore.command()
# @click.argument("scores", nargs=-1, type=click.Path())
# @click.option('--name-baidu', default='姓名', type=click.STRING)
# @click.option('--sid-baidu', default='学号', type=click.STRING)
# @click.option('--score-baidu', default='评分(0-100)', type=click.STRING)
# @click.option('--submitted', default='提交状态', type=click.STRING)
# @click.option('--time-submitted', default='学生提交时间', type=click.STRING)
# @click.option('--ranked', default='评阅状态', type=click.STRING)
# @click.option('--score-max', default=98, type=int)
# @click.option('--score-min', default=80, type=int)
# def rank(scores, **kwargs):
#     """Rank based submission order."""
#     fields = [kwargs['sid_baidu'], kwargs['score_baidu'],
#               kwargs['submitted'], kwargs['time_submitted'], kwargs['ranked']]
#     for xlsfile in scores:
#         data = load_xls(xlsfile)
#         header = data[0]
#         indices = find_column_index(header, fields)
#         col_sid, col_sc, col_sub, col_time, col_ranked = indices
#         data_rank, data_usub = split_ranking(data[1:], col_sub, col_ranked)
#         for row in data_usub:
#             row[col_sc] = 0
#         data_rank = time_based_rank(data_rank, col_sc, col_time,
#                                     kwargs['score_max'], kwargs['score_min'])
#         data = sorted(data_rank + data_usub,
#                       key=lambda row: row[col_sid])
#         write_xls([header] + data, xlsfile)

        
# @xduscore.command()
# @click.argument("scores", nargs=-1, type=click.Path())
# @click.option("-o", "--output", type=click.Path())
# @click.option("-w", "--weights", multiple=True, type=click.FLOAT)
# @click.option("-t", "--tasks", multiple=True, type=click.STRING)
# @click.option('--name-baidu', default='姓名', type=click.STRING)
# @click.option('--sid-baidu', default='学号', type=click.STRING)
# @click.option('--score-baidu', default='评分(0-100)', type=click.STRING)
# def collect(scores, weights, tasks, output, **kwargs):
#     """Collected scores downloaded from AI studio and save to output."""
#     assert len(tasks) == len(weights)
#     assert len(scores) == len(weights)

#     scores_all, students = {}, {}
#     fields = [kwargs["sid_baidu"], kwargs["score_baidu"], kwargs["name_baidu"]]
#     for xlsfile, task in zip(scores, tasks):
#         data = load_xls(xlsfile)
#         sid_col, score_col, name_col = find_column_index(data[0], fields)
#         score_task = value_dict(data, sid_col, score_col, True)
#         scores_all[task] = score_task
        
#         stus = value_dict(data, sid_col, name_col)
#         students = merge_students(students, stus)

#     weights_dict = {t: w for t, w in zip(tasks, weights)}
#     data, errors = merge_scores(scores_all, weights_dict)
#     fields = [kwargs["sid_baidu"]] + list(tasks) + ["加权平均分"]
#     form = score_dict_to_table(data, fields)
#     for row in filter(lambda x: x[-1] < 60, form[1:]):
#         print(row)
#     fmt = output.split(".")[-1]
#     if "csv" == fmt:
#         write_csv(form, output)
#     elif "xls" == fmt:
#         write_xls(form, output)

        
# @xduscore.command()
# @click.argument("forms", nargs=-1, type=click.Path(exists=True))
# @click.option("-f", "--file-src", type=click.Path())
# @click.option("-s", "--score-src", default="期末成绩", type=click.STRING)
# @click.option('--score-dest', default='期末成绩(数字)', type=click.STRING)
# @click.option('--score-rename', default=False, is_flag=True,
#               help="Rename destination score to source.")
# @click.option('--sid-src', default='学号(文本)', type=click.STRING)
# @click.option('--sid-dest', default='学号(文本)', type=click.STRING)
# def fill(forms, **kwargs):
#     """Fill forms for uploading to Xidian."""
#     scores = load_xls(kwargs['file_src'])
#     fields = [kwargs["sid_src"], kwargs['score_src']]
#     sid_col, score_col = find_column_index(scores[0], fields)
#     for row in scores:
#         if isinstance(row[sid_col], float):
#             row[sid_col] = f'{row[sid_col]:.0f}'
#     scores_dict = value_dict(scores, sid_col, score_col, True)

#     for xlsfile in forms:
#         score_form = load_xls(xlsfile)
#         sid_col, score_col = find_column_index(score_form[0], [kwargs["sid_dest"], kwargs["score_dest"]])
#         if kwargs['score_rename']:
#             score_form[0][score_col] = kwargs['score_src']
#         fill_score_form(score_form, scores_dict, sid_col, score_col)
#         write_xls(score_form, xlsfile)


# @xduscore.command()
# @click.argument("classes", nargs=-1, type=click.Path(exists=True))
# @click.option("-k", "--homework", nargs=1, type=click.Path(exists=True))
# @click.option('--sid-xdu', default='学号(文本)', type=click.STRING)
# def organize(classes, homework, **kwargs):
#     """Organize homeworks donwloaded from baidu for archival."""
#     fields = [kwargs["sid_xdu"]]
#     student_classes = {}
#     for cls in classes:
#         class_name = cls.split(".")[0]
#         if not os.path.exists(class_name):
#             os.mkdir(class_name)
#         # hw_path = os.path.join(class_name, homework)
#         # if not os.path.exists(hw_path):
#         #     os.mkdir(hw_path)
#         data = load_xls(cls)
#         sid_col, = find_column_index(data[0], fields)
#         for row in data[1:]:
#             sid = row[sid_col]
#             student_classes[sid] = class_name

#     for path, _, files in os.walk(homework):
#         for name in files:
#             if ".pdf" == name[-4:].lower():
#                 the_sid = name.split(".")[0].split("-")[-1]
#                 if the_sid in student_classes:
#                     the_cls = student_classes[the_sid]
#                     hw_src = os.path.join(path, name)
#                     hw_dst = os.path.join(the_cls, "", name)
#                     # print(f"Copying {hw_src} to {hw_dst}...")
#                     shutil.copyfile(hw_src, hw_dst)
#                 else:
#                     print(f"Cannot find the class for student {the_sid}.")
               

# @xduscore.command()
# @click.argument("scores", nargs=1, type=click.Path(exists=True))
# @click.option("-k", "--homework", nargs=1, type=click.Path(exists=True))
# @click.option('--name-baidu', default='姓名', type=click.STRING)
# @click.option('--sid-baidu', default='学号', type=click.STRING)
# @click.option('--score-baidu', default='评分(0-100)', type=click.STRING)
# def homework(scores, homework, **kwargs):
#     score_hw = {}
#     for folder in os.listdir(homework):
#         stu_folder = os.path.join(homework, folder)
#         files = os.listdir(stu_folder)
#         score_hw[folder] = 100

#     data = load_xls(scores)
#     fields = [kwargs["sid_baidu"], kwargs["score_baidu"], kwargs["name_baidu"]]
#     sid_col, score_col, name_col = find_column_index(data[0], fields)
#     for idx in range(1, len(data)):
#         sid = data[idx][sid_col]
#         data[idx][score_col] = score_hw.get(sid, 0)

#     write_xls(data, scores)
        

# @xduscore.command()
# @click.argument("accuracies", nargs=1, type=click.Path(exists=True))
# @click.argument("sheet", nargs=1, type=click.Path(exists=True))
# def challenge(accuracies, sheet, **kwargs):
#     """Assign scores based on challenge accuracies."""
#     with open(accuracies) as istream:
#         reader = csv.reader(istream)
#         header = next(reader)
#         scores = {}
#         for row in reader:
#             sid = row[0]
#             sc = round(float(row[1])*100)
#             if sc > 95:
#                 sc = 100
#             if sc < 60:
#                 sc = 60
#             scores[sid] = sc
#     # print("\n".join(scores.keys()))

#     data = load_xls(sheet)
#     fields = ["学号", "竞赛", "姓名"]
#     sid_col, score_col, name_col = find_column_index(data[0], fields)
#     for idx in range(1, len(data)):
#         sid = str(int(data[idx][sid_col]))
#         if sid in scores:
#             data[idx][score_col] = scores[sid]
#         else:
#             print(sid, data[idx][sid_col], data[idx][name_col])

#     write_xls(data, "AISE-2022-updated.xls")

# 
# score_helper.py ends here
